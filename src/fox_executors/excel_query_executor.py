#!/usr/bin/env python3
"""
Excel 本地数据问答原子执行器 (Fox-path Executor)
读取 .xlsx 文件，用本地 LLM 回答关于数据的问题。
"""
from __future__ import annotations

import os
import subprocess
import sys
from pathlib import Path

# 长表会撑爆 27B 的 prefill 时间，容易触发 Fox 子进程 180s 超时
_DEFAULT_MAX_ROWS = int(os.environ.get("EXCEL_QUERY_MAX_ROWS", "32"))
_MAX_CONTEXT_CHARS = int(os.environ.get("EXCEL_QUERY_MAX_CONTEXT_CHARS", "14000"))


def _read_excel_summary(excel_path: str, max_rows: int = _DEFAULT_MAX_ROWS) -> str:
    """读取 Excel 文件并返回文本摘要（表头 + 前N行数据）。"""
    try:
        import openpyxl
    except ImportError:
        # fallback: 用 subprocess 调 python3 -c
        return _read_excel_via_subprocess(excel_path, max_rows)

    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb.active

        rows_data = []
        max_row_cached = ws.max_row
        max_col_cached = ws.max_column
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                rows_data.append(f"... (共 {max_row_cached} 行，已截断到 {max_rows} 行)")
                break
            cells = [str(c) if c is not None else "" for c in row]
            rows_data.append(" | ".join(cells))

        wb.close()
        total_rows = max_row_cached or len(rows_data)
        total_cols = max_col_cached or 0

        header = f"文件: {excel_path}\n行数: {total_rows}, 列数: {total_cols}\n"
        return header + "\n".join(rows_data)
    except Exception as e:
        return f"读取 Excel 失败: {e}"


def _read_excel_via_subprocess(excel_path: str, max_rows: int) -> str:
    """通过子进程调用 python 读取 Excel（处理 venv 里没装 openpyxl 的情况）。"""
    safe_path = repr(str(Path(excel_path).expanduser().resolve()))
    code = f"""
import json, sys
try:
    import openpyxl
    wb = openpyxl.load_workbook({safe_path}, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    mr, mc = ws.max_row, ws.max_column
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= {max_rows}:
            rows.append("... (truncated)")
            break
        rows.append(" | ".join(str(c) if c is not None else "" for c in row))
    wb.close()
    print(f"Rows: {{mr}}, Cols: {{mc}}")
    print("\\n".join(rows))
except Exception as e:
    print(f"ERROR: {{e}}")
"""
    try:
        res = subprocess.run(
            [sys.executable, "-c", code],
            capture_output=True, text=True, timeout=10,
        )
        return (res.stdout or "").strip()
    except Exception as e:
        return f"子进程读取失败: {e}"


def _ask_llm(context: str, question: str) -> str:
    """调用本地 LLM 回答关于 Excel 数据的问题。"""
    try:
        from openai import OpenAI
    except ImportError:
        print("[excel_query] ❌ 缺少 openai 库", file=sys.stderr, flush=True)
        return ""

    base_url = os.environ.get("FOX_LOCAL_LLM_BASE_URL", "http://127.0.0.1:8000/v1")
    api_key = os.environ.get("FOX_LOCAL_LLM_API_KEY", "not-needed")

    try:
        client = OpenAI(base_url=base_url, api_key=api_key, timeout=600.0)
        model_id = os.environ.get("FOX_LOCAL_LLM_MODEL", "").strip()
        if not model_id:
            models = client.models.list()
            model_id = models.data[0].id if models.data else ""
        if not model_id:
            print("[excel_query] ❌ 无法获取本地模型", file=sys.stderr, flush=True)
            return ""

        completion = client.chat.completions.create(
            model=model_id,
            messages=[
                {
                    "role": "system",
                    "content": "你是数据分析助手。根据提供的 Excel 片段回答用户问题。用中文，简短给出数字与结论，不要铺垫或 Markdown 长文。",
                },
                {"role": "user", "content": f"以下是 Excel 数据（可能已截断）：\n{context}\n\n问题：{question}"},
            ],
            temperature=0.2,
            max_tokens=min(int(os.environ.get("EXCEL_QUERY_MAX_TOKENS", "640")), 8192),
        )
        return (completion.choices[0].message.content or "").strip()
    except Exception as e:
        print(f"[excel_query] ❌ LLM 调用失败: {e}", file=sys.stderr, flush=True)
        return ""


def run(excel_path: str, question: str) -> int:
    """读取 Excel 并回答问题。"""
    excel_path = str(Path(excel_path).expanduser().resolve())
    if not Path(excel_path).exists():
        print(f"[excel_query] ❌ 文件不存在: {excel_path}", file=sys.stderr, flush=True)
        return 1

    if not question:
        print("[excel_query] ❌ 缺少问题", file=sys.stderr, flush=True)
        return 1

    print(f"[excel_query] 📊 读取 Excel: {excel_path}", flush=True)
    max_rows = _DEFAULT_MAX_ROWS
    context = _read_excel_summary(excel_path, max_rows=max_rows)
    if context.startswith("读取") or context.startswith("ERROR") or context.startswith("子进程"):
        print(f"[excel_query] ❌ {context}", file=sys.stderr, flush=True)
        return 2

    if len(context) > _MAX_CONTEXT_CHARS:
        context = context[:_MAX_CONTEXT_CHARS] + f"\n\n[已截断到前 {_MAX_CONTEXT_CHARS} 字符，环境变量可调整 EXCEL_QUERY_MAX_CONTEXT_CHARS]"

    print(f"[excel_query] 🤔 回答问题: {question}", flush=True)
    answer = _ask_llm(context, question)
    if not answer:
        print("[excel_query] ❌ LLM 返回为空", file=sys.stderr, flush=True)
        return 2

    # 输出结果
    print(f"[excel_query] 💡 {answer}", flush=True)

    # 复制到剪贴板
    try:
        pbcopy = subprocess.Popen(["pbcopy"], stdin=subprocess.PIPE)
        pbcopy.communicate(answer.encode("utf-8"))
        print(f"[excel_query] 📋 答案已复制到剪贴板", flush=True)
    except Exception:
        pass

    return 0


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("用法: python excel_query_executor.py <excel路径> <问题>", file=sys.stderr)
        sys.exit(1)

    sys.exit(run(sys.argv[1], sys.argv[2]))
